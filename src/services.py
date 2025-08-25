# src/services.py
"""
Service Layer for the Media Tracker Bot.
This layer contains the core business logic and orchestrates the pipeline.
It decouples the API (main.py) from the agents.
"""

import os
import asyncio
import shutil
import logging
import re
import pandas as pd
import traceback

from collections import Counter
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, DefaultDict
from collections import defaultdict
from jose import jwt, JWTError
from jose.exceptions import ExpiredSignatureError
from fastapi.security import OAuth2PasswordBearer
from fastapi import HTTPException, Request, status
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


from .models import CompetitorReport, Article
from .agents import MediaTrackerTeam, AgentManager
from .configs import AppSettings
from .task_state import task_manager

# Attempt to import Celery task wrapper. If the `celery` package is not
# installed or the module cannot be imported, we fall back to using
# asyncio.create_task for background execution.
try:
    from .celery_worker import run_pipeline_task  # type: ignore[import]

    CELERY_AVAILABLE = True
except Exception:
    run_pipeline_task = None  # type: ignore[assignment]
    CELERY_AVAILABLE = False

logger = logging.getLogger(__name__)

SECRET_KEY = "your_secret_key"
ALGORITHM = "HS256"
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/login")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "120"))


def create_access_token(data: dict, expires_delta: timedelta | None = None):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + (
        expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    )
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(request: Request) -> str:
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, detail="Token missing"
        )

    token = auth_header.split(" ")[1]
    try:
        payload = jwt.decode(
            token, SECRET_KEY, algorithms=[ALGORITHM]
        )  # t·ª± verify 'exp'
        return payload.get("sub")
    except ExpiredSignatureError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, detail="Token expired"
        )
    except JWTError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token"
        )


class PipelineService:
    """
    Core pipeline logic for media tracking. This function handles crawling and report generation.
    It is intended to be called by background task workers.
    """

    def __init__(self, app_settings: AppSettings, user_email: str):
        self.settings = app_settings
        self.team: Optional[MediaTrackerTeam] = None
        self.status_by_source = {}
        self.user_email = user_email
        self.agent_manager = AgentManager.get_instance()

    def cancel_provider(self, session_id: str, provider: str) -> bool:
        try:
            ok = self.agent_manager.cancel_provider(
                session_id=session_id, provider=provider
            )
            if ok:
                import time

                t0 = time.monotonic()
                while time.monotonic() - t0 < 2.0:
                    if not getattr(self.agent_manager, "is_running", lambda *_: False)(
                        session_id, provider
                    ):
                        break
                    time.sleep(0.05)
            return ok
        except Exception:
            logging.exception(
                "cancel_provider failed (session=%s, provider=%s)", session_id, provider
            )
            return False

    def _source_key_of(self, s):
        ref = getattr(s, "reference_name", None)
        if ref:
            return str(ref).strip().lower()
        t = str(getattr(s, "type", "")).strip().lower()
        d = str(getattr(s, "domain", "")).strip().lower()
        return f"{t}|{d}"

    async def run_pipeline_logic(
        self,
        session_id: str,
        user_email: str,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        custom_keywords: Optional[Dict[str, List[str]]] = None,
        selected_sources: Optional[List[str]] = None,
    ) -> Optional[CompetitorReport]:
        logger.info(f"[{session_id}] Starting media tracking pipeline...")
        self.current_session_id = session_id

        session_config = self.settings.crawl_config.model_copy(deep=True)
        session_config.media_sources = self.settings.crawl_config.media_sources.copy()

        if custom_keywords:
            session_config.keywords = custom_keywords

        # NEW: l·ªçc theo selected_sources n·∫øu user ch·ªçn
        if selected_sources:
            sel = {str(x).strip().lower() for x in selected_sources}

            def _match(s):
                return self._source_key_of(s) in sel

            filtered = [s for s in session_config.media_sources if _match(s)]
            session_config.media_sources = filtered

        if start_date and end_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                session_config.date_range_days = (end_dt - start_dt).days
            except ValueError:
                logger.error(
                    f"[{session_id}] Invalid date format. Using default date range."
                )
                start_dt = None
                end_dt = None
        else:
            start_dt = None
            end_dt = None

        # L·∫•y l·∫°i task hi·ªán t·∫°i
        tasks = task_manager.get_tasks(user_email)
        task = next((t for t in tasks if t["session_id"] == session_id), None)
        articles_so_far = [Article(**a) for a in task.get("articles_so_far", [])]

        self.team = MediaTrackerTeam(
            config=session_config,
            start_date=start_dt,
            end_date=end_dt,
            session_id=session_id,
            user_email=user_email,
            on_progress_update=self._update_task_progress,
            check_cancelled=self.is_cancelled,
            check_paused=self.should_pause,
            articles_so_far=articles_so_far,
            source_status_list=task.get("source_status_list", []),
        )

        report = await asyncio.wait_for(
            self.team.run_full_pipeline(),
            timeout=self.settings.crawl_config.total_pipeline_timeout,
        )

        if report:
            await self._save_report(report, user_email)
            logger.info(f"[{session_id}] Pipeline completed successfully.")
        else:
            logger.warning(
                f"[{session_id}] Pipeline finished without generating a report."
            )
        return report

    def is_cancelled(self):
        tasks = task_manager.get_tasks(self.user_email)
        task = next(
            (t for t in tasks if t["session_id"] == self.current_session_id), None
        )
        return task and task["status"] == "cancelled"

    def should_pause(self):
        tasks = task_manager.get_tasks(self.user_email)
        task = next(
            (t for t in tasks if t["session_id"] == self.current_session_id), None
        )
        return task and task["status"] == "paused"

    def resume_task_worker(self, session_id):
        logger.info(f"[{session_id}] Resuming pipeline execution (resume_task_worker)")
        use_celery = CELERY_AVAILABLE and os.getenv("CELERY_ENABLED", "0") == "1"
        if use_celery:
            try:
                # Dispatch resumed execution to Celery
                run_pipeline_task.delay(
                    session_id=session_id,
                    user_email=self.user_email,
                    start_date=None,
                    end_date=None,
                    custom_keywords=None,
                    selected_sources=None,
                )
                return
            except Exception as exc:
                logger.exception(
                    f"Failed to dispatch Celery resume task for session {session_id}: {exc}"
                )
        # Fallback to local async task
        asyncio.create_task(self._task_worker(session_id))

    def _update_task_progress(
        self, source_name, completed, failed, progress, current_task
    ):
        all_sources = [s.name for s in self.settings.crawl_config.media_sources]

        current_tasks = task_manager.get_tasks(self.user_email)
        task = next(
            (t for t in current_tasks if t["session_id"] == self.current_session_id),
            None,
        )

        if task.get("status") not in ["pending", "running"]:
            return

        if task:
            # L·∫•y danh s√°ch source ƒë√£ completed t·ª´ team

            completed_or_failed = [
                s["source_name"]
                for s in self.team.source_status_list
                if s["status"] in ("completed", "failed")
            ]

            remaining = [s for s in all_sources if s not in completed_or_failed]

            task_manager.update_task(
                self.user_email,
                self.current_session_id,
                {
                    "current_source": source_name,
                    "completed_sources": completed,
                    "failed_sources": failed,
                    "progress": progress,
                    "current_task": current_task,
                    "status": "running",
                    "remaining_sources": remaining,
                    "source_status_list": self.team.source_status_list,  # L∆∞u lu√¥n v√†o task
                },
            )

    async def retry_task_worker(app_settings, user_email, session_id):
        new_service = PipelineService(app_settings, user_email)
        await new_service._task_worker(session_id)

    async def _task_worker(self, session_id):
        tasks = task_manager.get_tasks(self.user_email)
        task = next((t for t in tasks if t["session_id"] == session_id), None)

        if not task:
            logger.info(f"[{session_id}] Task not found, stopping worker.")
            return

        status = task["status"]

        if status == "paused":
            logger.info(f"[{session_id}] Task is paused. Waiting to resume...")
            # ƒê·ª£i cho ƒë·∫øn khi kh√¥ng c√≤n paused
            while True:
                current_status = task_manager.get_task_status(
                    self.user_email, session_id
                )
                if current_status == "paused":
                    await asyncio.sleep(2)
                elif current_status == "cancelled":
                    logger.info(
                        f"[{session_id}] Task was cancelled while paused. Exiting."
                    )
                    return  # üëâ D·ª´ng ngay, kh√¥ng ti·∫øp t·ª•c
                else:
                    break  # Resume h·ª£p l·ªá
            logger.info(
                f"[{session_id}] Resumed. Checking if cancelled before proceeding..."
            )

        # Ki·ªÉm tra l·∫°i tr·∫°ng th√°i tr∆∞·ªõc khi ti·∫øp t·ª•c
        current_status = task_manager.get_task_status(self.user_email, session_id)
        if current_status == "cancelled":
            logger.info(f"[{session_id}] Task is cancelled. Exiting worker.")
            return

        if current_status == "pending":
            logger.info(f"[{session_id}] Task is pending. Preparing pipeline setup...")
            task_manager.update_task(
                self.user_email,
                session_id,
                {
                    "status": "pending",
                    "current_task": "Preparing tools and resources...",
                    "progress": 0.0,
                },
            )
            await asyncio.sleep(5)  # Cho FE th·∫•y tr·∫°ng th√°i pending

            # Ki·ªÉm tra l·∫°i tr·∫°ng th√°i tr∆∞·ªõc khi chuy·ªÉn sang running
            current_status = task_manager.get_task_status(self.user_email, session_id)
            if current_status == "cancelled":
                logger.info(
                    f"[{session_id}] Task was cancelled during pending. Exiting."
                )
                return

            task_manager.update_task(
                self.user_email,
                session_id,
                {
                    "status": "running",
                    "current_task": "Crawling... please wait.",
                    "progress": 0.0,
                },
            )

        # B·∫Øt ƒë·∫ßu pipeline th·ª±c s·ª±
        try:
            logger.info(f"[{session_id}] Starting pipeline execution.")
            params = task["params"]

            await self.run_pipeline_logic(
                session_id,
                self.user_email,
                start_date=params["start_date"],
                end_date=params["end_date"],
                custom_keywords=params["custom_keywords"],
                selected_sources=params.get("selected_sources") or [],
            )
            task_manager.update_task(
                self.user_email, session_id, {"status": "completed", "progress": 100.0}
            )

        except asyncio.CancelledError:
            logger.warning(f"[{session_id}] ‚ö†Ô∏è Task forcefully cancelled.")
            task_manager.update_task(
                self.user_email,
                session_id,
                {
                    "status": "cancelled",
                    "current_task": "Pipeline was cancelled",
                },
            )

        except Exception as e:
            logger.error(
                f"[{session_id}] Pipeline failed: {e}\n{traceback.format_exc()}"
            )
            task["retry_count"] += 1
            if task["retry_count"] < task["max_retries"]:
                logger.info(
                    f"[{session_id}] Retry {task['retry_count']} / {task['max_retries']}"
                )
                task_manager.update_task(
                    self.user_email,
                    session_id,
                    {
                        "status": "pending",
                        "current_source": None,
                        "progress": 0.0,
                        "current_task": "Chu·∫©n b·ªã retry sau l·ªói",
                    },
                )
                await asyncio.sleep(5)
                asyncio.create_task(
                    self._task_worker(session_id)
                )  # T·ª± g·ªçi l·∫°i ƒë·ªÉ retry
            else:
                task_manager.update_task(
                    self.user_email, session_id, {"status": "failed", "error": str(e)}
                )

    def run_background_task(
        self,
        user_email,
        session_id,
        start_date=None,
        end_date=None,
        custom_keywords=None,
        selected_sources=None,
    ):
        if (
            selected_sources
            and isinstance(selected_sources, list)
            and len(selected_sources) > 0
        ):
            sel = {str(x).strip().lower() for x in selected_sources}

            def _match(s):
                return self._source_key_of(s) in sel

            media_sources = [
                s for s in self.settings.crawl_config.media_sources if _match(s)
            ]
        else:
            media_sources = self.settings.crawl_config.media_sources
        total_sources = len(media_sources)
        task_data = {
            "session_id": session_id,
            "status": "pending",
            "start_time": datetime.now().isoformat(),
            "progress": 0.0,
            "retry_count": 0,
            "max_retries": 2,
            # Th√¥ng tin c·∫ßn cho dashboard
            "params": {
                "start_date": start_date,
                "end_date": end_date,
                "custom_keywords": custom_keywords,
                "selected_sources": selected_sources or [],
            },
            "current_source": None,
            "total_sources": total_sources,
            "completed_sources": 0,
            "failed_sources": 0,
            "current_task": "Kh·ªüi ƒë·ªông pipeline",
            "source_status_list": [],
        }

        task_manager.add_task(self.user_email, task_data)
        task_manager.set_task_attr(
            self.user_email,
            session_id,
            "media_sources",
            [
                {
                    "source_key": self._source_key_of(s),
                    "name": s.name,
                    "type": s.type,
                    "domain": s.domain,
                    "reference_name": getattr(s, "reference_name", None),
                }
                for s in media_sources
            ],
        )

        # v√† l∆∞u ri√™ng danh s√°ch key (d√πng hi·ªÉn th·ªã/kh√¥i ph·ª•c)
        task_manager.set_task_attr(
            user_email,
            session_id,
            "selected_source_keys",
            [self._source_key_of(s) for s in media_sources],
        )

        # Choose between Celery and local async task execution
        use_celery = CELERY_AVAILABLE and os.getenv("CELERY_ENABLED", "0") == "1"
        if use_celery:
            # Dispatch to Celery worker; do not await
            try:
                run_pipeline_task.delay(
                    session_id=session_id,
                    user_email=user_email,
                    start_date=start_date,
                    end_date=end_date,
                    custom_keywords=custom_keywords,
                    selected_sources=selected_sources,
                )
            except Exception as exc:
                logger.exception(
                    f"Failed to dispatch Celery task for session {session_id}: {exc}"
                )
                # Fallback to local execution
                asyncio.create_task(self._task_worker(session_id))
        else:
            # Fall back to local asynchronous execution
            asyncio.create_task(self._task_worker(session_id))

        return session_id

    def get_status(self) -> Dict:
        """Gets the current status of the service and the running team."""
        return {"tasks": task_manager.get_tasks(self.user_email)}

    def _sanitize_user_name(self, username: str) -> str:
        """Tr·∫£ v·ªÅ t√™n th∆∞ m·ª•c an to√†n t·ª´ t√™n user."""
        username = username.strip().lower()
        sanitized = re.sub(r"[^a-zA-Z0-9]+", "_", username)
        return sanitized.strip("_") or "unknown_user"

    async def _save_report(self, report: CompetitorReport, user_email: str):
        """Saves the report to JSON and Excel files."""

        def _clean_one_line(s: str) -> str:
            s = re.sub(r"\s+", " ", str(s or "")).strip()
            return s

        def _dedup_preserve_order(items):
            seen = set()
            out = []
            for x in items:
                if not x:
                    continue
                if x not in seen:
                    seen.add(x)
                    out.append(x)
            return out

        # (tu·ª≥ nhu c·∫ßu) th·ª© t·ª± ng√†nh c·ªë ƒë·ªãnh theo template c·ªßa b·∫°n
        TEMPLATE_INDUSTRIES = [
            "D·∫ßu ƒÉn",
            "Gia v·ªã",
            "G·∫°o & Ng≈© c·ªëc",
            "S·ªØa (UHT)",
            "Baby Food",
            "Home Care",
        ]

        folder_name = self._sanitize_user_name(user_email)
        reports_dir = self.settings.reports_dir / folder_name
        reports_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        try:
            # ---------------- JSON ----------------
            json_file = reports_dir / f"report_{timestamp}.json"
            with open(json_file, "w", encoding="utf-8") as f:
                f.write(report.model_dump_json(indent=2))

            # ---------------- Build Overall (m·ªói ng√†nh = 1 d√≤ng, gi·ªØ T√äN ƒë·∫ßu b√°o) ----------------
            brands_by_ind = defaultdict(list)   # industry -> [brand, ...]
            papers_by_ind = defaultdict(set)    # industry -> {paper, ...}
            count_by_ind  = defaultdict(int)    # industry -> total articles

            for a in report.articles:
                ind = _clean_one_line(getattr(a, "nganh_hang", None))
                if not ind:
                    continue

                # brands
                for b in (getattr(a, "nhan_hang", None) or []):
                    b = _clean_one_line(b)
                    if b:
                        brands_by_ind[ind].append(b)

                # papers
                if getattr(a, "dau_bao", None):
                    papers_by_ind[ind].add(_clean_one_line(a.dau_bao))

                # count
                count_by_ind[ind] += 1

            rows = []
            # Tr∆∞·ªõc ti√™n ghi c√°c ng√†nh theo template (k·ªÉ c·∫£ kh√¥ng c√≥ b√†i)
            for ind in TEMPLATE_INDUSTRIES:
                brands = _dedup_preserve_order(brands_by_ind.get(ind, []))
                nhan_hang_txt = ", ".join(brands)  # 1 d√≤ng
                paper_names = _dedup_preserve_order(list(papers_by_ind.get(ind, set())))
                paper_names_txt = ", ".join(paper_names)  # 1 d√≤ng
                rows.append({
                    "Ng√†nh h√†ng": ind,
                    "Nh√£n h√†ng": nhan_hang_txt,
                    "C√°c ƒë·∫ßu b√°o": paper_names_txt,      # gi·ªØ T√äN
                    "S·ªë l∆∞·ª£ng b√†i": count_by_ind.get(ind, 0),
                })

            # Sau ƒë√≥ th√™m c√°c ng√†nh ph√°t sinh (n·∫øu c√≥) ngo√†i template
            # extra_inds = sorted(set(brands_by_ind.keys()) - set(TEMPLATE_INDUSTRIES))
            # for ind in extra_inds:
            #     brands = _dedup_preserve_order(brands_by_ind[ind])
            #     nhan_hang_txt = ", ".join(brands)
            #     paper_names = _dedup_preserve_order(list(papers_by_ind[ind]))
            #     paper_names_txt = ", ".join(paper_names)
            #     rows.append({
            #         "Ng√†nh h√†ng": ind,
            #         "Nh√£n h√†ng": nhan_hang_txt,
            #         "C√°c ƒë·∫ßu b√°o": paper_names_txt,
            #         "S·ªë l∆∞·ª£ng b√†i": count_by_ind[ind],
            #     })

            overall_df = pd.DataFrame(rows)

            # ---------------- EXCEL ----------------
            excel_file = reports_dir / f"report_{timestamp}.xlsx"
            with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
                # 1) Summary - Overall
                overall_df.to_excel(writer, sheet_name="Summary - Overall", index=False)

                ws = writer.sheets["Summary - Overall"]
                # ƒê·ªãnh d·∫°ng: h·∫°n ch·∫ø xu·ªëng d√≤ng, co ch·ªØ n·∫øu c·∫ßn
                ws.column_dimensions["A"].width = 18   # Ng√†nh h√†ng
                ws.column_dimensions["B"].width = 48   # Nh√£n h√†ng
                ws.column_dimensions["C"].width = 70   # C√°c ƒë·∫ßu b√°o (t√™n d√†i)
                ws.column_dimensions["D"].width = 14   # S·ªë l∆∞·ª£ng b√†i
                for r in range(2, ws.max_row + 1):
                    ws[f"A{r}"].alignment = Alignment(vertical="center")
                    ws[f"B{r}"].alignment = Alignment(wrap_text=False, vertical="center")
                    ws[f"C{r}"].alignment = Alignment(wrap_text=False, shrink_to_fit=True, vertical="center")
                    ws[f"D{r}"].alignment = Alignment(vertical="center")

                # 2) Summary theo ng√†nh (1 sheet / ng√†nh)
                for s in report.industry_summaries:
                    sheet_name = f"Summary - Ng√†nh {s.nganh_hang[:25]}"
                    rows_ind = []
                    for brand in s.nhan_hang:
                        # L·ªçc b√†i theo brand & ng√†nh
                        related_articles = [
                            article
                            for article in report.articles
                            if brand in (article.nhan_hang or [])
                            and article.nganh_hang == s.nganh_hang
                        ]
                        # C·ª•m n·ªôi dung unique, s·∫Øp x·∫øp ·ªïn ƒë·ªãnh
                        clusters = sorted({
                            _clean_one_line(article.cum_noi_dung)
                            for article in related_articles
                            if getattr(article, "cum_noi_dung", None)
                        })
                        cluster_text = "\n".join(
                            f"{i+1}. {c}" for i, c in enumerate(clusters)
                        ) if clusters else ""

                        rows_ind.append(
                            {
                                "Nh√£n h√†ng": _clean_one_line(brand),
                                "C·ª•m n·ªôi dung": cluster_text,
                                "S·ªë l∆∞·ª£ng b√†i": len(related_articles),
                            }
                        )

                    df_ind = pd.DataFrame(rows_ind)
                    df_ind.to_excel(writer, sheet_name=sheet_name, index=False)
                    ws_ind = writer.sheets[sheet_name]
                    ws_ind.column_dimensions["A"].width = 35
                    ws_ind.column_dimensions["B"].width = 70
                    ws_ind.column_dimensions["C"].width = 16
                    # wrap cho c·ªôt ‚ÄúC·ª•m n·ªôi dung‚Äù
                    for r in range(2, ws_ind.max_row + 1):
                        ws_ind[f"A{r}"].alignment = Alignment(wrap_text=False, vertical="center")
                        ws_ind[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
                        ws_ind[f"C{r}"].alignment = Alignment(vertical="center")

                # 3) M·ªói nh√£n h√†ng m·ªôt sheet b√†i b√°o (ƒë·∫∑t ng√†nh theo ƒëa s·ªë)
                articles_by_brand: DefaultDict[str, List[Article]] = defaultdict(list)
                articles_no_brand: List[Article] = []

                for article in report.articles:
                    if article.nhan_hang:
                        for b in article.nhan_hang:
                            articles_by_brand[b].append(article)
                    else:
                        articles_no_brand.append(article)

                for brand, articles in articles_by_brand.items():
                    # Ng√†nh theo ƒëa s·ªë
                    if articles:
                        ind_majority = Counter(
                            _clean_one_line(a.nganh_hang)
                            for a in articles
                            if getattr(a, "nganh_hang", None)
                        ).most_common(1)[0][0]
                    else:
                        ind_majority = ""

                    sheet_name = f"Ng√†nh {ind_majority} - {brand}"
                    sheet_name = sheet_name[:31]  # Excel limit

                    df_brand = pd.DataFrame(
                        [
                            {
                                "STT": a.stt,
                                "Ng√†y ph√°t h√†nh": a.ngay_phat_hanh.strftime("%d/%m/%Y"),
                                "ƒê·∫ßu b√°o": a.dau_bao,
                                "C·ª•m n·ªôi dung": a.cum_noi_dung_chi_tiet or a.cum_noi_dung,
                                "T√≥m t·∫Øt n·ªôi dung": a.tom_tat_noi_dung,
                                "Link b√†i b√°o": a.link_bai_bao,
                                "Keywords": ", ".join(a.keywords_found or []),
                            }
                            for a in articles
                        ]
                    )
                    df_brand["STT"] = range(1, len(df_brand) + 1)
                    df_brand = df_brand[["STT"] + [c for c in df_brand.columns if c != "STT"]]
                    df_brand.to_excel(writer, sheet_name=sheet_name, index=False)

                # 4) Sheet cho c√°c b√†i kh√¥ng c√≥ nh√£n h√†ng
                if articles_no_brand:
                    ind_first = _clean_one_line(articles_no_brand[0].nganh_hang)
                    sheet_name = f"General - {ind_first}"
                    sheet_name = sheet_name[:31]

                    df_gen = pd.DataFrame(
                        [
                            {
                                "STT": a.stt,
                                "Ng√†y ph√°t h√†nh": a.ngay_phat_hanh.strftime("%d/%m/%Y"),
                                "ƒê·∫ßu b√°o": a.dau_bao,
                                "C·ª•m n·ªôi dung": a.cum_noi_dung_chi_tiet or a.cum_noi_dung,
                                "T√≥m t·∫Øt n·ªôi dung": a.tom_tat_noi_dung,
                                "Link b√†i b√°o": a.link_bai_bao,
                            }
                            for a in articles_no_brand
                        ]
                    )
                    df_gen["STT"] = range(1, len(df_gen) + 1)
                    df_gen = df_gen[["STT"] + [c for c in df_gen.columns if c != "STT"]]
                    df_gen.to_excel(writer, sheet_name=sheet_name, index=False)

            # ---------------- latest symlink/copy ----------------
            for ext in ["json", "xlsx"]:
                latest_path = reports_dir / f"latest_report.{ext}"
                target_file = reports_dir / f"report_{timestamp}.{ext}"
                if latest_path.exists() or latest_path.is_symlink():
                    latest_path.unlink()
                shutil.copy(target_file, latest_path)

            logger.info(f"Report saved to {json_file} and {excel_file}")
        except Exception as e:
            logger.error(f"Failed to save report: {e}", exc_info=True)

